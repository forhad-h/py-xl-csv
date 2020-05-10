from openpyxl import load_workbook
from csv import writer, QUOTE_MINIMAL

file_name = 'vogtland-product-list.xlsx'
sheet_name = 'Height Adjustable Coil Overs'
product_name = sheet_name
image_url = '/wp-content/uploads/2020/03/RACE-SPRINGS.png'

wb = load_workbook(file_name)
sheet = wb[sheet_name]

# return the value


def v(s): return s.value


products_file = open(f'P-{sheet_name}.csv', mode='w')
query_info_file = open(f'Q-{sheet_name}.csv', mode="w")

product_writer = writer(
    products_file, delimiter=',', quotechar='"', quoting=QUOTE_MINIMAL)

query_info_writer = writer(
    query_info_file, delimiter=",", quotechar='*', quoting=QUOTE_MINIMAL)

product_writer.writerow([
    'Type',
    'Name',
    'Published',
    'Categories',
    'Visibility in catalog',
    'Short description',
    'In stock?',
    'Allow customer reviews?',
    'Sale price',
    'Regular price',
    'Images'
])

query_info_writer.writerow([
    'year',
    'make',
    'model',
    'product_image_url',
    'product_url'
])

for rn in range(sheet.max_row + 1):
    if (rn <= 1):
        continue
    models = [x.strip() for x in v(sheet[f'B{rn}']).split(',')]

    for model in models:

        make = v(sheet[f'A{rn}']).strip()
        model = model.strip()
        year = v(sheet[f'C{rn}']).strip()
        sale_price = v(sheet[f'H{rn}'])
        regular_price = v(sheet[f'G{rn}'])

        product_writer.writerow([
            'simple',
            f"{product_name} <span style='display: none;'>{make} {model} {year}</span>",
            1,
            v(sheet[f'A{rn}']),
            'visible',
            f"{make} {model} {year}",
            1,
            1,
            sale_price,
            regular_price,
            image_url
        ])

        product_url = '/product/' + \
            product_name.replace(' ', '-') + \
            '-' + make.replace(' ', '') + \
            '-' + model.replace(' ', '') + \
            '-' + year.replace(' ', '')

        query_info_writer.writerow([
            v(sheet[f'C{rn}']),
            v(sheet[f'A{rn}']),
            model,
            image_url,
            product_url.lower()
        ])

print("Done 😀")
